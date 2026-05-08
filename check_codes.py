#!/usr/bin/env python3
"""
Поиск владельцев кодов маркировки «Честный Знак».
Результат сохраняется в Excel-файл заданного формата.

Режимы:
  --public   Публичный API (без авторизации). Владелец НЕДОСТУПЕН.
  --true     True API (с токеном). Полная информация с владельцем.

True API endpoint:
  POST https://markirovka.crpt.ru/api/v3/true-api/cises/info?pg=<group>
  Header: Authorization: Bearer <token>
  Body: ["cis1", "cis2", ...]

Usage:
  python check_codes.py --true codes.txt -o result.xlsx
  python check_codes.py --public codes.txt -o result.xlsx
  python check_codes.py --true 0102901036818059215!bNNnpMl)pYo -o result.xlsx
"""

import json
import os
import sys
import time
import argparse
import urllib.request
import urllib.error
import textwrap
from datetime import datetime, timezone, timedelta
from pathlib import Path

# ── Конфигурация ────────────────────────────────────────────────────
PUBLIC_API = "https://mobile.api.crpt.ru/mobile/check"
TRUE_API   = "https://markirovka.crpt.ru/api/v3/true-api/cises/info"

TIMEOUT       = 20
RETRY         = 3
RETRY_DELAY   = 2

# Маппинг category → pg для True API
# В публичном API category = "lp", в True API pg может быть "lp", "light_industry" и т.д.
# Поэтому пробуем несколько вариантов
PG_ALIASES: dict[str, list[str]] = {
    "lp":           ["lp", "light_industry", "lightIndustry"],
    "milk":         ["milk", "dairy"],
    "tobacco":      ["tobacco"],
    "water":        ["water", "packed_water", "packedWater"],
    "beer":         ["beer", "brewery"],
    "shoes":        ["shoes", "footwear"],
    "perfume":      ["perfume", "perfumery"],
    "tires":        ["tires", "tyres"],
    "camera":       ["camera", "cameras"],
    "bicycle":      ["bicycle", "bicycles"],
    "furs":         ["furs"],
    "medicine":     ["medicine", "medicines", "drugs", "pharma"],
    "bio":          ["bio", "supplements", "dietary_supplements"],
    "antiseptic":   ["antiseptic", "antiseptics"],
    "wheelchair":   ["wheelchair", "wheelchairs"],
}

# Маппинг способа ввода в оборот
EMISSION_TYPE_MAP: dict[str, str] = {
    "PRODUCTION":          "Производство в РФ",
    "IMPORT":              "Ввезён в РФ",
    "REMAINS":             "Маркировка остатков",
    "REMARK":              "Перемаркировка",
    "CROSSBORDER":         "Трансграничная торговля",
    "COMMISSIONING":       "Ввод в оборот",
}

# Маппинг статусов (True API)
STATUS_MAP: dict[str, str] = {
    "EMITTED":             "Эмитирован",
    "APPLIED":             "Нанесён",
    "INTRODUCED":          "В обороте",
    "WRITTEN_OFF":         "Списан",
    "RETIRED":             "Выбыл",
    "DISAGGREGATED":       "Разагрегирован",
    "DESTROYED":           "Уничтожен",
    "SOLD":                "Продан",
    "WITHHELD":            "Приостановлен",
    "SHIPPED":             "Отгружен",
    "IN_CIRCULATION":      "В обороте",
    "IN_CIRCULATION_SOLD": "Продан",
    "WITHDRAWN":           "Выведен из оборота",
}

# ── Колонки Excel-файла ─────────────────────────────────────────────
EXCEL_HEADERS = [
    "Штрихкод",
    "GTIN",
    "Бренд",
    "Индекс картинки вида продукции",
    "Статус",
    "Количество кодов маркировки",
    "Владелец",
    "Производитель",
    "Дата ввода в оборот",
    "Способ ввода в оборот",
]


# ══════════════════════════════════════════════════════════════════════
# Загрузка .env
# ══════════════════════════════════════════════════════════════════════

def load_env(script_dir: Path) -> None:
    """Загружает переменные из .env, если он есть."""
    env_path = script_dir / ".env"
    if not env_path.exists():
        return
    try:
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, _, value = line.partition("=")
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and key not in os.environ:
                    os.environ[key] = value
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════════════
# HTTP-клиент
# ══════════════════════════════════════════════════════════════════════

def http_post(url: str, payload_str: str, headers: dict, debug: bool = False) -> tuple[int | None, str | None]:
    """POST с ретраями. Возвращает (http_code, body) или (None, None)."""
    for attempt in range(1, RETRY + 1):
        try:
            req = urllib.request.Request(
                url, data=payload_str.encode("utf-8"), headers=headers, method="POST",
            )
            with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
                body = resp.read().decode("utf-8")
                if debug:
                    print(f"\n  [DEBUG] URL: {url}", file=sys.stderr)
                    print(f"  [DEBUG] Headers: {json.dumps(dict(headers), ensure_ascii=False)}", file=sys.stderr)
                    print(f"  [DEBUG] Body sent: {payload_str}", file=sys.stderr)
                    print(f"  [DEBUG] HTTP {resp.status}, body: {body[:2000]}", file=sys.stderr)
                return (resp.status, body)
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="replace")
            if debug:
                print(f"\n  [DEBUG] URL: {url}", file=sys.stderr)
                print(f"  [DEBUG] Headers: {json.dumps(dict(headers), ensure_ascii=False)}", file=sys.stderr)
                print(f"  [DEBUG] Body sent: {payload_str}", file=sys.stderr)
                print(f"  [DEBUG] HTTP ERROR {e.code}, body: {body[:500]}", file=sys.stderr)
            return (e.code, body)
        except Exception as e:
            if debug:
                print(f"  [DEBUG] Exception (attempt {attempt}): {e}", file=sys.stderr)
            if attempt == RETRY:
                return (None, str(e))
            time.sleep(RETRY_DELAY * attempt)
    return (None, None)


# ══════════════════════════════════════════════════════════════════════
# Публичный API
# ══════════════════════════════════════════════════════════════════════

def public_check(code: str, debug: bool = False) -> dict | None:
    """Проверка одного кода через публичный API."""
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "okhttp/4.12.0",
    }
    status, body = http_post(PUBLIC_API, json.dumps({"code": code}), headers, debug=debug)
    if status and body:
        try:
            return json.loads(body)
        except json.JSONDecodeError:
            if debug:
                print(f"  [DEBUG] Ответ не JSON: {body[:200]}", file=sys.stderr)
    return None


# ══════════════════════════════════════════════════════════════════════
# True API
# ══════════════════════════════════════════════════════════════════════

def get_pg_from_public(data: dict) -> str:
    """Определяет товарную группу из ответа публичного API."""
    category = data.get("category") or ""
    if not category:
        category = (data.get("codeResolveData") or {}).get("codeCategory", "")
    return category


def resolve_pg_aliases(category: str) -> list[str]:
    """
    Возвращает список возможных значений pg для True API.
    Например, category='lp' → ['lp', 'light_industry', 'lightIndustry']
    """
    category_lower = category.lower()
    if category_lower in PG_ALIASES:
        return PG_ALIASES[category_lower]
    return [category_lower]


def true_check_batch(codes: list[str], pg: str, token: str, debug: bool = False,
                     log_fn=None) -> tuple[int | None, list[dict] | None]:
    """
    Пакетная проверка через True API.
    Возвращает (http_status, list_results) — при ошибке статус и None.
    log_fn: опциональная функция для логирования (message, tag).
    """
    url = f"{TRUE_API}?pg={pg}"
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Authorization": f"Bearer {token}",
    }

    status, body = http_post(url, json.dumps(codes), headers, debug=debug)

    if status == 401:
        msg = "❌ Токен недействителен (HTTP 401). Обновите токен."
        if body:
            try:
                err_msg = json.loads(body).get("error_message", "")
                if err_msg:
                    msg = f"❌ Токен недействителен (HTTP 401). Сервер: «{err_msg}»"
            except Exception:
                pass
        if log_fn:
            log_fn(msg, "error")
        else:
            print(msg, file=sys.stderr)
        return (status, None)
    if status == 403:
        msg = f"❌ Доступ запрещён (HTTP 403). Нет прав для товарной группы «{pg}»."
        if log_fn:
            log_fn(msg, "error")
        else:
            print(msg, file=sys.stderr)
        return (status, None)
    if status == 429:
        msg = "⚠ Слишком много запросов (HTTP 429). Превышен лимит — повторите позже."
        if log_fn:
            log_fn(msg, "warn")
        else:
            print(msg, file=sys.stderr)
        return (status, None)
    if status == 451:
        msg = "❌ Геоблокировка (HTTP 451). API доступен только с российских IP."
        if log_fn:
            log_fn(msg, "error")
        else:
            print(msg, file=sys.stderr)
        return (status, None)
    if status == 404:
        msg = f"⚠ HTTP 404 для pg={pg} — возможно, неверная товарная группа."
        if body and not debug:
            msg += f" Ответ: {body[:300]}"
        if log_fn:
            log_fn(msg, "warn")
        else:
            print(msg, file=sys.stderr)
        return (status, None)
    if 500 <= (status or 0) < 600:
        msg = f"❌ Ошибка сервера (HTTP {status}). Попробуйте позже."
        if log_fn:
            log_fn(msg, "error")
        else:
            print(msg, file=sys.stderr)
        return (status, None)
    if not status or not body:
        msg = "❌ Не удалось подключиться к API. Проверьте интернет и российский IP."
        if log_fn:
            log_fn(msg, "error")
        else:
            print(msg, file=sys.stderr)
        return (status, None)
    try:
        data = json.loads(body)
        return (status, data if isinstance(data, list) else [data])
    except json.JSONDecodeError:
        msg = f"⚠ Ответ не JSON: {body[:200]}"
        if log_fn:
            log_fn(msg, "warn")
        else:
            print(msg, file=sys.stderr)
        return (status, None)


def true_check_with_retry_pg(codes: list[str], category: str, token: str,
                             debug: bool = False, log_fn=None) -> tuple[int | None, list[dict] | None]:
    """
    Пробует все варианты pg для данной категории.
    Если первый вариант даёт 404 — пробует следующий.
    Возвращает (http_status, results).
    """
    aliases = resolve_pg_aliases(category)
    for pg in aliases:
        if debug:
            print(f"  [DEBUG] Пробую pg={pg}...", file=sys.stderr)
        status, data = true_check_batch(codes, pg, token, debug=debug, log_fn=log_fn)
        if data is not None:
            return (status, data)
        # Если 401 или 403 — нет смысла пробовать другие pg
        if status in (401, 403, 429, 451):
            return (status, None)
    return (None, None)


def true_check_auto(codes: list[str], token: str, debug: bool = False,
                    log_fn=None, stop_fn=None, batch_size: int = 100,
                    progress_fn=None) -> dict[str, dict]:
    """
    Определяет pg через публичный API, затем вызывает True API батчами.
    Возвращает {code: result_dict}.

    log_fn: функция логирования (message, tag) — для GUI
    stop_fn: функция, возвращающая True если нужно остановить
    batch_size: размер батча для True API запросов
    progress_fn: функция (done, total) для обновления прогресса
    """
    out: dict[str, dict] = {}

    def _log(msg: str, tag: str = "info"):
        if log_fn:
            log_fn(msg, tag)
        else:
            print(msg, file=sys.stderr)

    def _should_stop():
        return stop_fn() if stop_fn else False

    def _progress(done: int, total: int):
        if progress_fn:
            progress_fn(done, total)

    total_codes = len(codes)
    _log(f"📋 Шаг 1: определение товарных групп для {total_codes} кодов...")
    scanned = 0

    # Собираем category и публичные данные
    code_to_cat: dict[str, str] = {}
    public_results: dict[str, dict] = {}

    for code in codes:
        if _should_stop():
            _log("⏹ Остановлено на этапе определения товарных групп.", "warn")
            return out
        data = public_check(code, debug=debug)
        scanned += 1
        _progress(scanned, total_codes)
        if data:
            cat = get_pg_from_public(data)
            code_to_cat[code] = cat
            public_results[code] = data
            if debug:
                print(f"    {code[:30]}... → category={cat}", file=sys.stderr)
        else:
            out[code] = {"error": "Публичный API недоступен"}
            if debug:
                print(f"    {code[:30]}... → ⚠ нет данных", file=sys.stderr)
        time.sleep(0.1)

    if not code_to_cat:
        _log("❌ Не удалось определить товарную группу ни для одного кода.", "error")
        return out

    # Группируем по category
    cat_to_codes: dict[str, list[str]] = {}
    for code, cat in code_to_cat.items():
        cat_to_codes.setdefault(cat, []).append(code)

    # Шаг 2: батчи по batch_size через True API
    _log(f"📡 Шаг 2: запрос True API батчами по {batch_size} кодов...")
    total_to_check = len(code_to_cat)
    checked = 0

    for cat, group in cat_to_codes.items():
        pg_options = resolve_pg_aliases(cat)
        _log(f"  📦 Товарная группа «{cat}»: {len(group)} кодов (pg варианты: {pg_options})", "info")

        # Разбиваем группу на батчи
        for batch_start in range(0, len(group), batch_size):
            if _should_stop():
                _log("⏹ Остановлено на этапе проверки True API.", "warn")
                return out

            batch = group[batch_start:batch_start + batch_size]
            batch_num = batch_start // batch_size + 1
            total_batches = (len(group) + batch_size - 1) // batch_size
            _log(f"    Батч {batch_num}/{total_batches} ({len(batch)} кодов)...", "info")

            status, data = true_check_with_retry_pg(batch, cat, token, debug=debug, log_fn=log_fn)

            if data:
                found = set()
                for item in data:
                    cis_info = item.get("cisInfo", item)
                    c = cis_info.get("requestedCis", cis_info.get("cis", ""))
                    if c:
                        found.add(c)
                    out[c] = item
                for code in batch:
                    if code not in found:
                        out[code] = out.get(code, {})
                        if "error" not in out[code]:
                            out[code]["_public_data"] = public_results.get(code)
                checked += len(batch)
                _log(f"    ✓ Батч {batch_num}/{total_batches}: получено {len(data)} результатов", "success")
            else:
                error_msg = explain_http_status(status)
                for code in batch:
                    out[code] = out.get(code, {"error": f"True API: {error_msg}"})
                checked += len(batch)
                _log(f"    ✗ Батч {batch_num}/{total_batches}: {error_msg}", "error")

            _progress(scanned + checked, total_codes)
            time.sleep(0.15)

    return out


def explain_http_status(status: int | None) -> str:
    """Расшифровка HTTP-статуса для отображения в ошибках."""
    if status is None:
        return "нет соединения с сервером"
    mapping = {
        400: "некорректный запрос",
        401: "токен недействителен или просрочен",
        403: "доступ запрещён",
        404: "код не найден или неверная товарная группа",
        429: "превышен лимит запросов",
        451: "геоблокировка — нужен российский IP",
    }
    if status in mapping:
        return f"HTTP {status}: {mapping[status]}"
    if 500 <= status < 600:
        return f"HTTP {status}: ошибка сервера"
    return f"HTTP {status}"


# ══════════════════════════════════════════════════════════════════════
# Парсинг результата в строку Excel
# ══════════════════════════════════════════════════════════════════════

def extract_group_data(data: dict) -> dict | None:
    """Извлекает данные товарной группы (lpData, milkData...)."""
    for key in [
        "lpData", "milkData", "tobaccoData", "waterData", "beerData",
        "shoesData", "perfumeData", "tiresData", "cameraData",
        "bicycleData", "fursData", "wheelchairData",
    ]:
        if key in data:
            return data[key]
    return None


def iso_to_datetime_str(iso_str: str | None) -> str:
    """ISO-строка '2025-05-27T02:33:13.644Z' → '27.05.2025 05:33:13' (МСК)."""
    if not iso_str:
        return ""
    try:
        # Убираем Z и парсим
        s = iso_str.replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
        dt_msk = dt.astimezone(timezone(timedelta(hours=3)))
        return dt_msk.strftime("%d.%m.%Y %H:%M:%S")
    except (ValueError, TypeError):
        return iso_str


def ts_to_datetime_str(ts: int | str | None) -> str:
    """Timestamp (ms) → строка даты DD.MM.YYYY HH:MM:SS (МСК)."""
    if ts is None:
        return ""
    try:
        ts = int(ts)
        if ts > 1_000_000_000_000:
            ts = ts // 1000
        dt = datetime.fromtimestamp(ts, tz=timezone.utc) + timedelta(hours=3)
        return dt.strftime("%d.%m.%Y %H:%M:%S")
    except (ValueError, OSError):
        return str(ts) if ts else ""


def translate_status(status: str) -> str:
    """Переводит статус на русский."""
    return STATUS_MAP.get(status.upper(), status)


def translate_emission_type(emission_type: str) -> str:
    """Переводит способ ввода в оборот."""
    return EMISSION_TYPE_MAP.get(emission_type.upper(), emission_type)


def parse_result(code: str, data: dict, mode: str) -> list[str]:
    """
    Преобразует ответ API в строку для Excel (10 колонок).
    mode: "public" | "true"
    """
    if "error" in data:
        return [code, "", "", "", f"ОШИБКА: {data['error']}", "", "", "", "", ""]

    row = [""] * 10
    row[0] = code

    if mode == "true" and "_public_data" in data:
        return parse_public_row(code, data["_public_data"])

    if mode == "true":
        # ── True API: данные внутри cisInfo ──
        cis_info = data.get("cisInfo", data)
        cis_status = cis_info.get("status", cis_info.get("cisStatus", ""))
        owner_name  = cis_info.get("ownerName", "")
        owner_inn   = cis_info.get("ownerInn", "")
        producer    = cis_info.get("producerName", "")
        introduced  = cis_info.get("introducedDate", "")
        emission    = cis_info.get("emissionType", "")
        product     = cis_info.get("productName", "")
        gtin        = cis_info.get("gtin", "")
        brand       = cis_info.get("brand", "")

        row[1] = gtin
        row[2] = brand

        row[4] = translate_status(cis_status)
        row[5] = "1"

        parts = []
        if owner_name:
            parts.append(owner_name)
        if owner_inn:
            parts.append(f"ИНН {owner_inn}")
        row[6] = ", ".join(parts) if parts else ""

        row[7] = producer
        row[8] = iso_to_datetime_str(introduced)
        row[9] = translate_emission_type(emission)

        for field in ["imageIndex", "image_index", "pictureIndex"]:
            val = cis_info.get(field)
            if val is not None:
                row[3] = str(val)
                break
    else:
        return parse_public_row(code, data)

    return row


def parse_public_row(code: str, data: dict) -> list[str]:
    """Парсит строку из публичного API (10 колонок)."""
    row = [""] * 10
    row[0] = code

    pub_status = data.get("status", "")
    outer = data.get("outerStatus", "")
    status_str = outer or pub_status

    # GTIN из codeResolveData
    crd = data.get("codeResolveData") or {}
    if crd:
        row[1] = crd.get("gtin", "")

    # Brand из catalogData или productProperty
    catalog = (data.get("catalogData") or [None])[0]
    if catalog:
        row[2] = catalog.get("brand_name", "")

    if pub_status == "wrong":
        row[4] = "Невалидный код"
    elif data.get("codeFounded") is False:
        row[4] = "Не найден"
    else:
        row[4] = translate_status(status_str)

    row[5] = "1"

    group = extract_group_data(data)
    if group:
        # Brand может быть и в group-данных
        if not row[2]:
            row[2] = group.get("brand", "")
        # GTIN может быть в group.codeData
        if not row[1]:
            cd = group.get("codeData") or {}
            if cd:
                row[1] = cd.get("gtin", "")
        row[7] = group.get("producerName", "")
        row[8] = ts_to_datetime_str(group.get("introducedDate", ""))
        row[9] = translate_emission_type(
            (group.get("productProperty") or {}).get("emissionType", "")
        )

    if catalog and not row[7]:
        row[7] = catalog.get("producer_name", "")

    return row


# ══════════════════════════════════════════════════════════════════════
# Excel-экспорт
# ══════════════════════════════════════════════════════════════════════

def save_excel(rows: list[list[str]], output_path: str, title: str | None = None) -> None:
    """Сохраняет строки в Excel-файл."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ Нужен openpyxl: pip install openpyxl", file=sys.stderr)
        import csv
        csv_path = output_path.replace(".xlsx", ".csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(EXCEL_HEADERS)
            w.writerows(rows)
        print(f"✓ Сохранён CSV: {csv_path}")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    if title:
        ws.title = title[:31]

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    col_widths = [35, 20, 25, 15, 18, 12, 35, 30, 22, 22]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    if rows:
        ws.auto_filter.ref = f"A1:J{len(rows) + 1}"

    wb.save(output_path)
    print(f"\n✓ Результат сохранён: {output_path}")
    print(f"  Строк данных: {len(rows)}")


# ══════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════

def main():
    script_dir = Path(__file__).resolve().parent
    load_env(script_dir)

    parser = argparse.ArgumentParser(
        description="Проверка кодов маркировки «Честный Знак» с выгрузкой в Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""\
            Примеры:
              python check_codes.py --true codes.txt -o result.xlsx
              python check_codes.py --true --pg lp codes.txt -o result.xlsx
              python check_codes.py --public codes.txt -o result.xlsx
              python check_codes.py --true 0102901036818059215!bNNnpMl)pYo --debug

            Токен: в .env → CHESTNYZNAK_TOKEN=...
              или export CHESTNYZNAK_TOKEN="..."
              или --token "..."
        """),
    )
    parser.add_argument("codes", nargs="*", metavar="CODE")
    parser.add_argument("--true", action="store_true")
    parser.add_argument("--public", action="store_true")
    parser.add_argument("-f", "--file", help="Файл со списком кодов")
    parser.add_argument("--stdin", action="store_true")
    parser.add_argument("-o", "--output", default="result.xlsx")
    parser.add_argument("--delay", type=float, default=0.3)
    parser.add_argument("--token", help="Токен True API")
    parser.add_argument("--pg", help="Код товарной группы вручную (lp, milk, ...)")
    parser.add_argument("-v", "--verbose", action="store_true")
    parser.add_argument(
        "--debug", action="store_true",
        help="Показать URL запросов, заголовки, тело и полный ответ API"
    )

    args = parser.parse_args()

    use_true = args.true or not args.public

    # Сбор кодов
    codes = list(args.codes)
    if args.file:
        p = Path(args.file)
        if not p.exists():
            print(f"❌ Файл не найден: {args.file}", file=sys.stderr)
            sys.exit(1)
        codes.extend(l.strip() for l in p.read_text("utf-8").splitlines() if l.strip())
    if args.stdin:
        codes.extend(l.strip() for l in sys.stdin.read().splitlines() if l.strip())
    if not codes:
        parser.print_help()
        sys.exit(1)

    seen = set()
    unique = []
    for c in codes:
        if c not in seen:
            seen.add(c)
            unique.append(c)
    codes = unique

    mode_label = "True API" if use_true else "Публичный API"
    print(f"Режим: {mode_label}")
    print(f"Кодов для проверки: {len(codes)}")
    if args.debug:
        print(f"Режим отладки: ВКЛ (показываю URL, заголовки, тело запросов)")
    print()

    rows: list[list[str]] = []

    if use_true:
        token = args.token or os.environ.get("CHESTNYZNAK_TOKEN", "")
        if not token:
            print("❌ Токен не задан!", file=sys.stderr)
            print("   Укажи в .env: CHESTNYZNAK_TOKEN=...", file=sys.stderr)
            sys.exit(1)

        # Если указан --pg и один код — покажем URL для дебага
        if args.debug and args.pg:
            url = f"{TRUE_API}?pg={args.pg}"
            print(f"  [DEBUG] Будет отправлен запрос:", file=sys.stderr)
            print(f"    URL:     POST {url}", file=sys.stderr)
            print(f"    Headers: Authorization: Bearer <скрыто>", file=sys.stderr)
            print(f"             Content-Type: application/json", file=sys.stderr)
            print(f"    Body:    {json.dumps(codes[:3])}{'...' if len(codes) > 3 else ''}", file=sys.stderr)
            print(file=sys.stderr)

        if args.pg:
            print(f"Запрос True API (pg={args.pg}, {len(codes)} кодов)...")
            status, results = true_check_batch(codes, args.pg, token, debug=args.debug)
            if results:
                result_map = {}
                for item in results:
                    cis = item.get("cis", item.get("code", item.get("requestedCis", "")))
                    if cis:
                        result_map[cis] = item
                for code in codes:
                    item = result_map.get(code)
                    if item:
                        if args.verbose:
                            print(json.dumps(item, ensure_ascii=False, indent=2))
                        rows.append(parse_result(code, item, "true"))
                    else:
                        rows.append([code, "", "", "", "Нет данных в ответе", "", "", "", "", ""])
                print(f"  Получено результатов: {len(results)}/{len(codes)}")
            else:
                err_msg = explain_http_status(status)
                print(f"  ❌ Ошибка: {err_msg}", file=sys.stderr)
                for code in codes:
                    rows.append([code, "", "", "", f"ОШИБКА: True API — {err_msg}", "", "", "", "", ""])
        else:
            results = true_check_auto(codes, token, debug=args.debug)
            for code in codes:
                item = results.get(code, {"error": "Нет данных"})
                if args.verbose and "error" not in item:
                    print(json.dumps(item, ensure_ascii=False, indent=2))
                rows.append(parse_result(code, item, "true"))

    else:
        for i, code in enumerate(codes, 1):
            print(f"[{i}/{len(codes)}] {code[:40]}...", end=" ", flush=True)
            data = public_check(code, debug=args.debug)
            if data:
                if args.verbose:
                    print()
                    print(json.dumps(data, ensure_ascii=False, indent=2))
                else:
                    status = data.get("outerStatus") or data.get("status", "?")
                    product = data.get("productName", "")
                    print(f"→ {status} | {product[:50]}")
                rows.append(parse_result(code, data, "public"))
            else:
                print("→ ❌ нет данных")
                rows.append([code, "", "", "", "ОШИБКА: нет ответа от API", "", "", "", "", ""])
            if i < len(codes):
                time.sleep(args.delay)

    output_path = args.output
    if not output_path.startswith("/"):
        output_path = str(script_dir / output_path)
    save_excel(rows, output_path)

    ok = sum(1 for r in rows if not r[4].startswith("ОШИБКА"))
    errors = sum(1 for r in rows if r[4].startswith("ОШИБКА"))
    with_owner = sum(1 for r in rows if r[6] and r[6].strip())
    print(f"\n{'='*50}")
    print(f"СВОДКА: проверено {len(codes)} кодов")
    print(f"  Успешно: {ok}")
    if use_true:
        print(f"  Найдены владельцы: {with_owner}")
    print(f"  Ошибки: {errors}")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
