"""
Excel-экспорт результатов проверки.
"""
from __future__ import annotations

import sys

from .constants import EXCEL_HEADERS


def save_excel(rows: list[list[str]], output_path: str, title: str | None = None) -> None:
    """Сохраняет строки в Excel-файл с Catppuccin-стилем."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ Нужен openpyxl: pip install openpyxl", file=sys.stderr)
        import csv
        csv_path = output_path.replace(".xlsx", ".csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(EXCEL_HEADERS)
            w.writerows(rows)
        print(f"✓ Сохранён CSV: {csv_path}")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    if title:
        ws.title = title[:31]

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    col_widths = [35, 20, 25, 15, 18, 12, 35, 30, 22, 22]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    if rows:
        ws.auto_filter.ref = f"A1:J{len(rows) + 1}"

    wb.save(output_path)
    print(f"\n✓ Результат сохранён: {output_path}")
    print(f"  Строк данных: {len(rows)}")
